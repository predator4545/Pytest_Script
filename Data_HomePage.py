import openpyxl


class HomePage_Data:
    Login_Creds = {"Name":"rahul","Email":"shetty@gmail.com","Password":"34567890"}

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:/Users/Selvakumar/Desktop/Pranesh/PythonDataDriverDemo.xlsx")
        sheet = book.active  #pics the active sheet in the excel

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value  # loading excel data's into dictionary

        return[Dict]


#We can import a class from another file and access its class variables in another class without inheriting it.
# In Python, class variables are accessible via the class name directly,
# so there's no need to inherit from the other class unless required for design purposes.

#Key Points
#Direct Access via Class Name:

#In Python, class variables are shared among all instances of a class and can be accessed directly using the class name (ClassA.shared_variable).
#No Inheritance Required:

#Since class variables belong to the class itself (not the instance), there’s no need to inherit the class to access its class variables.
#Importing Across Files:

#You can import the class from another module (file) and directly access its class variables without needing to instantiate it.

#We can even Modify or Use the Variable in another class
#to access class variables, no need for object
#but to call any method(non static method) in a class, we need to create object for the class

#If we create static methods, then we can use without object creation, no need to have self parameter
